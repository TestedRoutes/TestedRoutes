#!/usr/bin/env python3
"""Import the Fiji country master workbook into places.yaml.

The master workbook (TestedRoutes_Fiji_Master_v<N>.xlsx) is the founder-owned
registry of every pin: identity, geo, per-SKU role/day/order triples, booking
link and printed go-slug. It is already a relational model wearing an Excel
costume; this script formalises it into content/countries/fiji/places.yaml so
publish-sku.mjs can load it into the private Postgres store.

Ownership rules, which matter more than the parsing:
  - The workbook is founder-edited and READ-ONLY to this script. The append-only
    versioning convention lives on the workbook (v1..vN); we always read the
    highest version present.
  - places.yaml is MACHINE-OWNED: regenerated wholesale on every run, git is its
    history. Never hand-edit it; edit the workbook and re-run
    `npm run import:master`. (This is the documented exception to the
    append-only rule, which protects founder-editable deliverables - the
    editable artifact here is the workbook.)

Validation posture: fail loudly, never guess.
  - The header row must match the generator's spec exactly
    (map-build/build_fiji_master.py HEADERS). A v11 with reshuffled columns
    must stop this script, not silently mis-map fields.
  - Every go-slug is checked against live Sanity affiliateLink docs with an
    anonymous GROQ read (the dataset is public; that is why no token is
    needed). Aliases are frozen contracts once printed - an unknown alias in
    the workbook is either a typo or an unregistered link, and both need a
    human. --skip-golink-check exists for offline runs.
"""

import argparse
import hashlib
import json
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
FIJI_DIR = ROOT / "content" / "countries" / "fiji"
OUT_PATH = FIJI_DIR / "places.yaml"
COUNTRY = "fj"
SHEET = "Fiji Master"
GUIDES = ["1D", "7D", "10D", "14D"]

EXPECTED_HEADERS = ["PinID", "Name", "Type", "Region", "Tier"]
for g in GUIDES:
    EXPECTED_HEADERS += [f"{g} Map role", f"{g} Day", f"{g} Order"]
EXPECTED_HEADERS += [
    "Description",
    "Latitude",
    "Longitude",
    "Map URL",
    "Category",
    "Booking link (destination)",
    "QR shortlink (go/)",
    "Internal notes (never exported)",
]

# The map-system convention names four roles, but the workbook has a fifth:
# "Service" marks rows that belong to a SKU without map placement (airlines,
# car rental, visa info, and the companion-map QR pseudo-pins, which the
# generator writes with null coordinates).
ROLES = {"route", "extra", "detour", "skip", "service"}

SANITY_QUERY_URL = (
    "https://y3gc8dx6.api.sanity.io/v2024-01-01/data/query/production?query="
    + urllib.parse.quote('*[_type == "affiliateLink"].slug.current')
)


def newest_master():
    candidates = []
    for p in FIJI_DIR.glob("TestedRoutes_Fiji_Master_v*.xlsx"):
        m = re.search(r"_v(\d+)\.xlsx$", p.name)
        if m:
            candidates.append((int(m.group(1)), p))
    if not candidates:
        sys.exit(f"No TestedRoutes_Fiji_Master_v<N>.xlsx found in {FIJI_DIR}")
    return max(candidates)[1]


def cell(v):
    """Normalise a workbook cell: strip strings, map ''/'N/A' to None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "N/A"):
            return None
    return v


def coord(v, pin_id, field, warnings):
    v = cell(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    # The generator writes "[DECIDE]" for unresolved pins; a founder edit could
    # leave other strings. Either way this pin has no usable coordinate yet.
    warnings.append(f"{pin_id}: {field} is {v!r}, storing null")
    return None


def go_slug(v, pin_id, problems):
    v = cell(v)
    if v is None:
        return None
    # Cell format is "testedroutes.com/go/<slug>" per the generator; accept a
    # bare slug too in case a hand-added row wrote it short.
    m = re.search(r"(?:testedroutes\.com)?/?go/([a-z0-9-]+)$", v)
    if m:
        return m.group(1)
    if re.fullmatch(r"[a-z0-9-]+", v):
        return v
    problems.append(f"{pin_id}: unparseable QR shortlink cell {v!r}")
    return None


def live_affiliate_slugs():
    with urllib.request.urlopen(SANITY_QUERY_URL, timeout=30) as r:
        data = json.load(r)
    return {s for s in data.get("result", []) if s}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-golink-check", action="store_true",
                    help="skip validating go-slugs against live Sanity (offline runs)")
    ap.add_argument("--strict-golinks", action="store_true",
                    help="treat unregistered go-slugs as errors instead of warnings")
    args = ap.parse_args()

    src = newest_master()
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"{src.name}: sheet {SHEET!r} not found (has {wb.sheetnames})")
    ws = wb[SHEET]
    rows = ws.iter_rows(values_only=True)

    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    if header[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        for i, (got, want) in enumerate(zip(header, EXPECTED_HEADERS)):
            if got != want:
                sys.exit(
                    f"{src.name}: header mismatch at column {i + 1}: "
                    f"got {got!r}, expected {want!r}. Refusing to guess - "
                    f"update EXPECTED_HEADERS if the workbook format changed on purpose."
                )
        sys.exit(f"{src.name}: header shorter than expected ({len(header)} cols)")

    places, problems, warnings = [], [], []
    seen = set()
    for row in rows:
        if row is None or cell(row[0]) is None:
            continue
        pin_id = str(cell(row[0]))
        if pin_id in seen:
            problems.append(f"duplicate PinID {pin_id}")
            continue
        seen.add(pin_id)

        skus = {}
        for gi, g in enumerate(GUIDES):
            role = cell(row[5 + gi * 3])
            day = cell(row[6 + gi * 3])
            order = cell(row[7 + gi * 3])
            if role is None:
                continue
            role_n = str(role).strip().lower()
            if role_n not in ROLES:
                problems.append(f"{pin_id}: unknown {g} Map role {role!r}")
                continue
            entry = {"role": role_n}
            if day is not None:
                entry["day"] = int(day)
            if order is not None:
                entry["order"] = int(order)
            skus[g] = entry

        place = {
            "pin_id": pin_id,
            "name": cell(row[1]),
            "type": cell(row[2]),
            "region": cell(row[3]),
            "tier": str(cell(row[4])) if cell(row[4]) is not None else None,
            "description": cell(row[17]),
            "lat": coord(row[18], pin_id, "Latitude", warnings),
            "lng": coord(row[19], pin_id, "Longitude", warnings),
            "map_url": cell(row[20]) if cell(row[20]) != "[DECIDE]" else None,
            "category": cell(row[21]),
            "booking_url": cell(row[22]),
            "go_slug": go_slug(row[23], pin_id, problems),
            "internal_notes": cell(row[24]),
            "skus": skus,
        }
        if not place["name"]:
            problems.append(f"{pin_id}: empty Name")
        places.append(place)

    if not args.skip_golink_check:
        try:
            live = live_affiliate_slugs()
        except Exception as e:  # noqa: BLE001 - any network failure means we can't validate
            sys.exit(f"go-link validation failed to reach Sanity ({e}); "
                     f"re-run with --skip-golink-check to import anyway")
        unknown = sorted(
            p["pin_id"] + " -> " + p["go_slug"]
            for p in places
            if p["go_slug"] and p["go_slug"] not in live
        )
        if unknown:
            # Unregistered slugs are a real workbook state, not necessarily a
            # typo: sync_go_links.py registers rows selectively, and a slug is
            # only a live contract once a deck prints its QR. So: warn loudly
            # by default (they must be registered before printing), fail only
            # under --strict-golinks.
            msg = ("go-slugs not registered as affiliateLink docs "
                   "(register with sync_go_links.py --live before any deck "
                   "prints them):\n    " + "\n    ".join(unknown))
            if args.strict_golinks:
                problems.append(msg)
            else:
                warnings.append(msg)

    for w in warnings:
        print(f"  warn: {w}")
    if problems:
        print(f"\n{src.name} failed import validation:", file=sys.stderr)
        for p in problems:
            print(f"  - {p}", file=sys.stderr)
        sys.exit(1)

    sha = hashlib.sha256(src.read_bytes()).hexdigest()[:16]
    header_comment = (
        f"# GENERATED by scripts/import_fiji_master.py from {src.name} "
        f"(sha256:{sha})\n"
        f"# Do not hand-edit: edit the workbook and re-run `npm run import:master`.\n"
        f"# The workbook is founder-owned and read-only to the importer.\n"
    )
    body = yaml.safe_dump(
        {"country": COUNTRY, "source": src.name, "places": places},
        sort_keys=False,
        allow_unicode=True,
        width=100,
        default_flow_style=False,
    )
    OUT_PATH.write_text(header_comment + body, encoding="utf-8", newline="\n")
    linked = sum(1 for p in places if p["go_slug"])
    print(f"{src.name}: {len(places)} places -> {OUT_PATH.relative_to(ROOT)} "
          f"({linked} with go-links)")


if __name__ == "__main__":
    main()
