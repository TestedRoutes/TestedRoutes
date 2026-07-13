"""Rebuild the TestedRoutes Action Tracker:
- Weeks 1-2 sheet: Paulius technical tasks only
- Weeks 3-4 sheet: Paulius technical tasks only
- Social Media sheet: ALL Jorune tasks (newsletter/social/youtube/content) from W1-W4
- Topic (2 words) column added
- Done tasks at the top of each section
- Brand swept across all sheets (pikelistravel.com → testedroutes.com, PikelisTravel → TestedRoutes)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import datetime

SRC = 'C:/Users/pauli/Downloads/20260425_TestedRoutes_Action_Tracker (1).xlsx'
DST = 'C:/Users/pauli/Downloads/20260427_TestedRoutes_Action_Tracker.xlsx'
shutil.copy(SRC, DST)

wb = load_workbook(DST)

# ==========================================================================
# 1. Read tasks
# ==========================================================================
def read_week1():
    ws = wb['Week 1']
    out = []
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(r, 4).value
        task = ws.cell(r, 5).value
        if not (task or topic):
            continue
        out.append({
            'week': 1, 'num': ws.cell(r, 2).value, 'day': ws.cell(r, 3).value,
            'topic': topic, 'task': task or topic, 'desc': ws.cell(r, 6).value,
            'owner': ws.cell(r, 7).value, 'priority': ws.cell(r, 8).value,
            'status': ws.cell(r, 9).value, 'start': ws.cell(r, 10).value,
            'due': ws.cell(r, 11).value, 'hrs_est': ws.cell(r, 12).value,
            'notes': ws.cell(r, 16).value,
        })
    return out

def read_week234(sheet, week_num):
    ws = wb[sheet]
    out = []
    for r in range(5, ws.max_row + 1):
        task = ws.cell(r, 4).value
        status = ws.cell(r, 8).value
        s_task = str(task or '').lower()
        if not task:
            continue
        if 'superseded' in s_task or str(status or '').lower() == 'superseded':
            continue
        out.append({
            'week': week_num, 'num': ws.cell(r, 2).value, 'day': ws.cell(r, 3).value,
            'topic': None, 'task': task, 'desc': ws.cell(r, 5).value,
            'owner': ws.cell(r, 6).value, 'priority': ws.cell(r, 7).value,
            'status': status, 'start': ws.cell(r, 9).value,
            'due': ws.cell(r, 10).value, 'hrs_est': ws.cell(r, 11).value,
            'notes': ws.cell(r, 15).value,
        })
    return out

# ==========================================================================
# 2. Owner classification (for "Founder" / unset)
# ==========================================================================
JORUNE_KW = [
    'reel', 'youtube', 'newsletter', 'beehiiv', 'social manager',
    'video editor', 'brand designer', 'designer brief', 'pinterest',
    'photo curation', 'film ', 'filming', 'tiktok', 'instagram',
    'social bios', 'design direction', 'design review',
    'editor brief', 'design concepts', 'logo + colour',
    'guide pdf template', 'ig post template', 'thumbnail template',
    'brand final review', 'guide template review', 'travel companions',
    'stories ', 'fb index', 'fb archive', 'welcome letter',
    'sender email', "sender display", 'profile pictures', 'profile photo',
    'pasidaryti', 'jorune', 'reach out to 10 travel newsletters',
    'list 20 target newsletters', 'feedback email',
]

def classify(task):
    t = (task or '').lower()
    for kw in JORUNE_KW:
        if kw in t:
            return 'Jorune'
    return 'Paulius'

def assign_owner(item):
    o = item.get('owner')
    if o in ('Paulius', 'Jorune'):
        return o
    return classify(item['task'])

# ==========================================================================
# 3. Topic (2 words) derivation
# ==========================================================================
TOPIC_RULES = [
    ('lock brand|brand decision|brand name', 'Brand'),
    ('pronunciation', 'Brand'),
    ('register social handle', 'Social handles'),
    ('register .com|domain ownership|domains', 'Domains'),
    ('trademark', 'Trademark'),
    ('business bank', 'Banking'),
    ('swiss ag|estonian|legal entity', 'Legal entity'),
    ('lawyer|privacy|gdpr|terms|legal sign', 'Legal'),
    ('polar|stripe|lemonsqueezy|currency|vat', 'Payments'),
    ('refund', 'Refunds'),
    ('analytics|plausible|posthog', 'Analytics'),
    ('search console|sitemap|seo', 'SEO'),
    ('affiliate|amazon|booking|gyg|getyourguide|viator|tiqets|skyscanner|impact|awin|safetywing|geniuslink|cj ', 'Affiliates'),
    ('utm|tracking', 'Tracking'),
    ('/go/|link cloak|cloudflare country', 'Link routing'),
    ('region-mapping', 'Affiliates'),
    ('site skeleton|kick off site|next.js|vercel|staging|production|publish.*guide', 'Site build'),
    ('landing page|guide product page|hub page', 'Site build'),
    ('apply.*brand design|new brand design across', 'Brand rollout'),
    ('test /go|test every affiliate', 'QA'),
    ('qa ', 'QA'),
    ('checkout flow', 'Checkout'),
    ('stripe.*test|stripe.*live|stripe checkout', 'Checkout'),
    ('beta website ux|alpha website ux', 'UX testing'),
    ('focus group', 'UX testing'),
    ('zoom calls with tech', 'UX testing'),
    ('ai prompt|ai polish', 'AI prompts'),
    ('guide template|draft.*guide template', 'Guide templates'),
    ('write 5 more guides|draft 3 guides', 'Guides'),
    ('embed affiliate', 'Guides'),
    ('founder selects|creator builds', 'Guides'),
    ('photo curation|photo library|lightroom', 'Photos'),
    ('fb.*data|fb.*archive|fb html|fb parser|facebook data', 'FB archive'),
    ('reel', 'Reels'),
    ('youtube', 'YouTube'),
    ('newsletter|beehiiv', 'Newsletter'),
    ('welcome letter|welcome email', 'Newsletter'),
    ('sender email|sender domain', 'Newsletter'),
    ('film batch|film first|film youtube', 'Filming'),
    ('script.*youtube', 'YouTube'),
    ('video editor', 'Video editor'),
    ('brand designer|design.*concept|review concepts|lock logo|finalise brand designer', 'Brand design'),
    ('design direction|design review', 'Brand design'),
    ('pinterest', 'Pinterest'),
    ('social manager', 'Social manager'),
    ('social bios|ig.*bio|tiktok bio', 'Social bios'),
    ('post.*reel|post first reel|reels.*published|reels cadence|schedule.*reel', 'Reels'),
    ('post brand designer brief|post video editor brief', 'Hiring'),
    ('shortlist|interview', 'Hiring'),
    ('story 1|story 2|stories 1|stories 3|stories 6|stories 10|stories 13|stories 17|scan fb index|polish.*publish', 'Stories'),
    ('soft.launch|monitor first', 'Launch'),
    ('reply to every|send 10 free guides', 'Launch'),
    ('reddit value drop', 'Reddit'),
    ('hand-pick.*testimonial|testimonial', 'Testimonials'),
    ('week.*review|month.*retrospective|update business plan|month 2', 'Review'),
    ('synthesize.*review|review feedback', 'Review'),
    ('reach out|partnerships|swap', 'Partnerships'),
    ('email structure|email forwarding', 'Email setup'),
    ('2fa|2-factor|2 factor|password', 'Security'),
    ('accounting|taxes|treuh|bookkeeping', 'Accounting'),
    ('official address|domizilagentur', 'Address'),
    ('reconfirm.*team|brothers|wife|family', 'Team'),
    ('20 destinations|first 20 destination', 'Destinations'),
    ('repurpose pikelistravel|pikelistravel.com.*redirect', 'Domain'),
    ('migrate sanity', 'Sanity'),
    ('price test', 'Pricing'),
    ('returns made simple|refund.*copy', 'Refunds'),
    ('feedback email', 'Feedback'),
    ('review business case', 'Business case'),
]

import re
def derive_topic(item):
    if item.get('topic'):
        return item['topic']
    t = (item.get('task') or '').lower()
    for pattern, label in TOPIC_RULES:
        if re.search(pattern, t):
            return label
    # Default: take first 1-2 meaningful words
    words = re.findall(r"[A-Za-z]+", item.get('task') or '')
    if words:
        return ' '.join(words[:2]).title()
    return ''

# ==========================================================================
# 4. Today's-work updates
# ==========================================================================
STATUS_UPDATES = {
    'site skeleton live on staging': 'Done',
    'kick off site build': 'Done',
}
NOTE_UPDATES = {
    'site skeleton live on staging': 'Done — testedroutes.com is live with homepage, /guides, /inspire, /destinations/switzerland, /about (PR #10 merged 2026-04-26)',
    'set up refunds@pikelistravel.com': 'Update target email to refunds@testedroutes.com (pikelistravel.com being repurposed)',
    'open business bank account': 'Blocked — pending Swiss AG vs Estonian OU decision',
}

NEW_TASKS_PAULIUS = [
    {
        'week': 1, 'topic': 'Domain',
        'task': 'Repurpose pikelistravel.com — set up HTTPS-capable redirect or new use',
        'desc': 'pikelistravel.com is fully detached from TestedRoutes infra. Namecheap URL forwarder works HTTP-only; HSTS-cached browsers fail HTTPS handshake. Decide path forward (Cloudflare on a new business-email account, redirect-only Vercel project, or accept HTTP-only).',
        'priority': 'P2', 'status': 'Not started',
        'due': datetime.datetime(2026, 4, 27),
        'notes': 'Deferred from 2026-04-26 session. Domain repurpose target ~2026-05-26.',
    },
    {
        'week': 1, 'topic': 'Sanity',
        'task': 'Migrate Sanity ownership to paulius@testedroutes.com',
        'desc': 'Both gmail + work email are added but neither is org owner. Sanity gates ownership-transfer support behind Enterprise / Growth add-on. Plan: ask in their Discord community.',
        'priority': 'P1', 'status': 'Blocked',
        'notes': 'Account migration off personal gmail — the only service still pending.',
    },
]
NEW_TASKS_JORUNE = [
    {
        'week': 1, 'topic': 'Newsletter',
        'task': 'Beehiiv: write/update welcome email + set sender domain',
        'desc': 'Welcome email and sender display refresh now that brand = TestedRoutes. Custom sending domain (newsletter@testedroutes.com) requires DKIM + return-path setup; coordinate with Paulius on SPF (single-row gotcha — modify existing in place).',
        'priority': 'P0', 'status': 'Not started',
        'due': datetime.datetime(2026, 4, 27),
        'notes': 'Deferred from 2026-04-26 session.',
    },
]

# ==========================================================================
# 5. Build dataset
# ==========================================================================
all_w12 = read_week1() + read_week234('Week 2', 2)
all_w34 = read_week234('Week 3', 3) + read_week234('Week 4', 4)

def apply_updates(items):
    for it in items:
        # Owner
        it['owner_final'] = assign_owner(it)
        # Topic
        it['topic'] = derive_topic(it)
        # Status / note overrides
        tk = (it['task'] or '').lower()
        for kw, new_status in STATUS_UPDATES.items():
            if kw in tk:
                it['status'] = new_status
        for kw, new_note in NOTE_UPDATES.items():
            if kw in tk:
                existing = it.get('notes')
                it['notes'] = (str(existing) + ' | ' if existing else '') + new_note
        # Brand swap in text
        for k in ('task', 'desc', 'notes', 'topic'):
            v = it.get(k)
            if isinstance(v, str):
                it[k] = (v.replace('pikelistravel.com', 'testedroutes.com')
                         .replace('PikelisTravel', 'TestedRoutes')
                         .replace('Pikelis Travel', 'TestedRoutes'))

apply_updates(all_w12)
apply_updates(all_w34)

# Append new tasks
for t in NEW_TASKS_PAULIUS:
    t['owner_final'] = 'Paulius'
    t['day'] = None; t['hrs_est'] = None; t['start'] = None
    t['num'] = None
    all_w12.append(t)
for t in NEW_TASKS_JORUNE:
    t['owner_final'] = 'Jorune'
    t['day'] = None; t['hrs_est'] = None; t['start'] = None
    t['num'] = None
    all_w12.append(t)

# Sort: Done at top, then Blocked, In progress, Not started; then priority; then week; then due date
STATUS_ORDER = {'Done': 0, 'Blocked': 1, 'In progress': 2, 'Not started': 3, None: 3, '': 3}
PRIO_ORDER = {'P0': 0, 'P1': 1, 'P2': 2, None: 3, '': 3, '—': 3}

def sort_key(t):
    return (
        STATUS_ORDER.get(t.get('status') or 'Not started', 3),
        PRIO_ORDER.get(t.get('priority'), 3),
        t.get('week') or 0,
        t.get('due') or datetime.datetime(2099, 1, 1),
    )

# ==========================================================================
# 6. Sweep brand strings across non-Week sheets
# ==========================================================================
SHEETS_TO_SWEEP = ['Website Build', 'Decisions', 'Guides Pipeline', 'Content Calendar',
                   'Hiring', 'README', 'Dashboard', 'Daily Log', 'Expenses', 'Site Map',
                   'Guide Templates', 'Stories', 'Feedback Rounds', 'Affiliate Programs',
                   'Refunds Log', 'Reviews & Testimonials', 'Regional Affiliates',
                   'Advisor Outreach']

for sn in SHEETS_TO_SWEEP:
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str):
                new_v = (v.replace('pikelistravel.com', 'testedroutes.com')
                          .replace('PikelisTravel', 'TestedRoutes')
                          .replace('Pikelis Travel', 'TestedRoutes')
                          .replace('mantas@testedroutes.com', 'refunds@testedroutes.com'))
                if new_v != v:
                    c.value = new_v

# Specific updates to Decisions sheet — mark brand decision as done
if 'Decisions' in wb.sheetnames:
    ws = wb['Decisions']
    # Find the brand-name row and add note that it's decided
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and 'brand name: keep' in v.lower():
                # Add status marker in nearby columns if there's a status column
                ws.cell(r, c).value = v + ' — DECIDED: TestedRoutes (2026-04-26)'

# ==========================================================================
# 7. Sheet writers
# ==========================================================================
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=12)
TITLE_FONT = Font(name='Arial', bold=True, size=14)
SUBTITLE_FONT = Font(name='Arial', italic=True, color='555555', size=10)
DEFAULT_FONT = Font(name='Arial', size=10)

PAULIUS_FILL = PatternFill('solid', start_color='1F4E79')
JORUNE_FILL = PatternFill('solid', start_color='8B4513')
HEADER_FILL = PatternFill('solid', start_color='666666')
TOPIC_FILL = PatternFill('solid', start_color='F2F2F2')

STATUS_FILLS = {
    'Done': PatternFill('solid', start_color='C6EFCE'),
    'In progress': PatternFill('solid', start_color='FFEB9C'),
    'Blocked': PatternFill('solid', start_color='FFC7CE'),
    'Not started': PatternFill('solid', start_color='FFFFFF'),
}

THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ['#', 'Wk', 'Topic', 'Task', 'Description', 'Priority', 'Status', 'Due date', 'Hrs est.', 'Notes']
COL_WIDTHS = [5, 5, 16, 50, 60, 9, 13, 12, 9, 50]

def write_header_row(ws, r):
    for ci, label in enumerate(COLS, 1):
        c = ws.cell(r, ci)
        c.value = label
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[r].height = 24

def write_data_row(ws, r, n, it):
    ws.cell(r, 1).value = n
    ws.cell(r, 2).value = it.get('week')
    ws.cell(r, 3).value = it.get('topic')
    ws.cell(r, 4).value = it.get('task')
    ws.cell(r, 5).value = it.get('desc')
    ws.cell(r, 6).value = it.get('priority')
    ws.cell(r, 7).value = it.get('status') or 'Not started'
    ws.cell(r, 8).value = it.get('due')
    ws.cell(r, 9).value = it.get('hrs_est')
    ws.cell(r, 10).value = it.get('notes')

    for ci in range(1, len(COLS) + 1):
        c = ws.cell(r, ci)
        c.font = DEFAULT_FONT
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = BORDER

    st = it.get('status') or 'Not started'
    ws.cell(r, 7).fill = STATUS_FILLS.get(st, STATUS_FILLS['Not started'])
    ws.cell(r, 7).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
    ws.cell(r, 3).fill = TOPIC_FILL
    ws.cell(r, 3).font = Font(name='Arial', size=10, bold=True)

    if isinstance(ws.cell(r, 8).value, datetime.datetime):
        ws.cell(r, 8).number_format = 'yyyy-mm-dd'

    ws.cell(r, 1).alignment = Alignment(vertical='top', horizontal='center')
    ws.cell(r, 2).alignment = Alignment(vertical='top', horizontal='center')
    ws.cell(r, 6).alignment = Alignment(vertical='top', horizontal='center')

def build_paulius_sheet(name, label, date_range, items):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    paulius = sorted([i for i in items if i['owner_final'] == 'Paulius'], key=sort_key)
    counts = {}
    for it in paulius:
        s = it.get('status') or 'Not started'
        counts[s] = counts.get(s, 0) + 1

    ws.cell(1, 1).value = label
    ws.cell(1, 1).font = TITLE_FONT
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 26
    summary = f"{date_range}    —    {len(paulius)} tasks  ({', '.join(f'{k}: {v}' for k, v in sorted(counts.items()))})"
    ws.cell(2, 1).value = summary
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.merge_cells('A2:J2')

    write_header_row(ws, 4)
    for n, it in enumerate(paulius, 1):
        write_data_row(ws, 4 + n, n, it)
    ws.freeze_panes = 'A5'

def build_social_sheet(name, label, items_w12, items_w34):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    j_w12 = sorted([i for i in items_w12 if i['owner_final'] == 'Jorune'], key=sort_key)
    j_w34 = sorted([i for i in items_w34 if i['owner_final'] == 'Jorune'], key=sort_key)

    counts = {}
    for it in j_w12 + j_w34:
        s = it.get('status') or 'Not started'
        counts[s] = counts.get(s, 0) + 1

    ws.cell(1, 1).value = label
    ws.cell(1, 1).font = TITLE_FONT
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 26
    summary = f"All Jorune tasks (newsletter, social, YouTube, content) Weeks 1-4    —    {len(j_w12) + len(j_w34)} tasks  ({', '.join(f'{k}: {v}' for k, v in sorted(counts.items()))})"
    ws.cell(2, 1).value = summary
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.merge_cells('A2:J2')

    next_row = 4
    # Weeks 1-2 section
    ws.cell(next_row, 1).value = f'WEEKS 1-2 — Mon 4 May – Sun 17 May ({len(j_w12)} tasks)'
    ws.cell(next_row, 1).font = SECTION_FONT
    ws.cell(next_row, 1).fill = JORUNE_FILL
    ws.cell(next_row, 1).alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[next_row].height = 24
    ws.merge_cells(start_row=next_row, end_row=next_row, start_column=1, end_column=len(COLS))
    next_row += 1
    write_header_row(ws, next_row)
    next_row += 1
    for n, it in enumerate(j_w12, 1):
        write_data_row(ws, next_row, n, it)
        next_row += 1
    next_row += 1

    # Weeks 3-4 section
    ws.cell(next_row, 1).value = f'WEEKS 3-4 — Mon 18 May – Sun 31 May ({len(j_w34)} tasks)'
    ws.cell(next_row, 1).font = SECTION_FONT
    ws.cell(next_row, 1).fill = JORUNE_FILL
    ws.cell(next_row, 1).alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[next_row].height = 24
    ws.merge_cells(start_row=next_row, end_row=next_row, start_column=1, end_column=len(COLS))
    next_row += 1
    write_header_row(ws, next_row)
    next_row += 1
    for n, it in enumerate(j_w34, 1):
        write_data_row(ws, next_row, n, it)
        next_row += 1

    ws.freeze_panes = 'A5'

# Build the three new sheets
build_paulius_sheet('Weeks 1-2', 'Weeks 1-2 — Paulius (Technical)',
                    'Mon 4 May – Sun 17 May  |  Foundation + Momentum', all_w12)
build_paulius_sheet('Weeks 3-4', 'Weeks 3-4 — Paulius (Technical)',
                    'Mon 18 May – Sun 31 May  |  Acceleration + Soft Launch', all_w34)
build_social_sheet('Social Media', 'Social Media — Jorune (Content / Newsletter / YouTube)',
                   all_w12, all_w34)

# Move new sheets to the front
order = ['Weeks 1-2', 'Weeks 3-4', 'Social Media'] + [s for s in wb.sheetnames if s not in ('Weeks 1-2', 'Weeks 3-4', 'Social Media')]
wb._sheets = [wb[s] for s in order]

# Rename old weeks for clarity
for old, new in [('Week 1', 'Week 1 (orig)'), ('Week 2', 'Week 2 (orig)'),
                 ('Week 3', 'Week 3 (orig)'), ('Week 4', 'Week 4 (orig)')]:
    if old in wb.sheetnames and new not in wb.sheetnames:
        wb[old].title = new

wb.save(DST)
print(f'Saved: {DST}')
print(f'Sheet order: {wb.sheetnames[:6]}...')

p_w12 = sum(1 for i in all_w12 if i['owner_final'] == 'Paulius')
p_w34 = sum(1 for i in all_w34 if i['owner_final'] == 'Paulius')
j_total = sum(1 for i in all_w12 + all_w34 if i['owner_final'] == 'Jorune')
print(f'Weeks 1-2 (Paulius): {p_w12}')
print(f'Weeks 3-4 (Paulius): {p_w34}')
print(f'Social Media (Jorune): {j_total}')
