#!/usr/bin/env python3
"""
content_status.py - reconcile the three sources of truth for Inspire content.

  folders  content/countries/{country}/inspire/{ID}   = what exists
  Sanity   story documents                             = what is live
  Excel    Content Plan "Inspire" sheet Status column  = what is planned

Nobody maintains status by hand: this script REPORTS drift between the three
(the same pre-flight role sync_go_links.py plays for QR codes). Run it before
any launch.

Usage:
  python scripts/content_status.py --plan "<Content-Plan.xlsx>"          # report
  python scripts/content_status.py --plan "<...>" --write-status        # also fix
                                    the Status column where reality disagrees

--write-status only touches two transitions: "-> Published" when a planned row
is live in Sanity, and "-> Drafted" when a row says Published but is NOT live.
It never invents rows and never changes any other column.
"""
import argparse
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parents[1]
COUNTRIES = REPO / "content" / "countries"
LEDGER = COUNTRIES / "publish-ledger.json"
API_VERSION = "2026-04-24"


def load_env(path=REPO / ".env.local"):
    if not path.exists():
        sys.exit(".env.local not found - run from the repo")
    for line in open(path, encoding="utf-8"):
        m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)=(.*)$", line)
        if m:
            os.environ.setdefault(m.group(1), m.group(2).strip().strip('"'))


def sanity_slugs():
    project = os.environ["NEXT_PUBLIC_SANITY_PROJECT_ID"]
    dataset = os.environ.get("NEXT_PUBLIC_SANITY_DATASET", "production")
    q = urllib.parse.quote('*[_type=="story"]{"slug": slug.current, language, storyId}')
    url = f"https://{project}.api.sanity.io/v{API_VERSION}/data/query/{dataset}?query={q}"
    with urllib.request.urlopen(url, timeout=60) as r:
        return json.loads(r.read())["result"]


def scan_folders():
    """{story_id: country} for every story folder on disk."""
    out = {}
    for country in COUNTRIES.iterdir():
        if not country.is_dir() or country.name.startswith("_"):
            continue
        insp = country / "inspire"
        roots = [insp] if insp.is_dir() else [country]
        for root in roots:
            for d in root.iterdir():
                if d.is_dir() and d.name not in ("photos", "generated"):
                    out[d.name] = country.name
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--plan", required=True, help="Content Plan xlsx")
    ap.add_argument("--write-status", action="store_true")
    a = ap.parse_args()

    load_env()
    ledger = json.loads(LEDGER.read_text(encoding="utf-8")) if LEDGER.exists() else {}
    live = sanity_slugs()
    live_slugs = {s["slug"] for s in live if s.get("slug")}
    folders = scan_folders()

    wb = load_workbook(a.plan)
    ws = wb["Inspire"]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    ix = {h: hdr.index(h) for h in hdr if h}
    id_c, st_c = ix["ID"], ix["Status"]

    drift = {"live_not_marked": [], "marked_not_live": [], "planned_no_folder": [],
             "no_ledger_but_live": 0}
    planned_ids = set()
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        sid = str(row[id_c].value or "").strip()
        if not sid:
            continue
        planned_ids.add(sid)
        status = str(row[st_c].value or "").strip()
        slug = (ledger.get(sid) or {}).get("slug") or re.sub(r"^\d{4}_", "", sid)
        is_live = slug in live_slugs
        if is_live and status != "Published":
            drift["live_not_marked"].append(f"{sid}  (Status: {status or 'blank'})")
            if a.write_status:
                row[st_c].value = "Published"; fixed += 1
        if not is_live and status == "Published":
            drift["marked_not_live"].append(sid)
            if a.write_status:
                row[st_c].value = "Drafted"; fixed += 1
        if sid not in folders and status not in ("Published",):
            drift["planned_no_folder"].append(sid)

    unplanned = sorted(set(folders) - planned_ids)

    print(f"folders: {len(folders)}   plan rows: {len(planned_ids)}   live in Sanity: {len(live_slugs)}\n")
    def block(title, items):
        print(f"{title}: {len(items)}")
        for i in items[:25]:
            print(f"   {i}")
        if len(items) > 25:
            print(f"   ... +{len(items)-25} more")
        print()
    block("LIVE in Sanity but plan does not say Published", drift["live_not_marked"])
    block("Plan says Published but NOT live in Sanity", drift["marked_not_live"])
    block("Planned (not yet published) with NO folder on disk", drift["planned_no_folder"])
    block("Folder on disk but not in the plan", unplanned)

    if a.write_status and fixed:
        wb.save(a.plan)
        print(f"Status column updated on {fixed} row(s).")
    elif drift["live_not_marked"] or drift["marked_not_live"]:
        print("Re-run with --write-status to sync the Status column to reality.")


if __name__ == "__main__":
    main()
