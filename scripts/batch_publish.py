#!/usr/bin/env python3
"""
TestedRoutes - batch publisher for Inspire stories.

Reads the Content Plan masterfile, and for every row whose Status is
"To be published":
  1. locates content/inspire/{country}/{ID}/
  2. extracts {ID}_story*.docx -> generated/{ID}_story_en.md
  3. converts photos/ masters -> generated/{ID}_*.jpg (<=2560px, q85)
  4. cuts every video -> generated/{stem}.mp4 (7s, 720p, muted), keeping the
     _{n}- slot in the name; stale renditions of renamed/removed masters are
     cleaned from generated/ so reordering is a pure rename + re-run
  5. drafts {ID}_meta_en.yaml (plan fields + Claude-proposed fields)
  6. runs scripts/publish-to-sanity.mjs on the folder
  7. writes the ledger + review CSV, and sets the row's Status to
     "Published" in the masterfile.

Rows with any other status are never touched.

Usage:
  python scripts/batch_publish.py --plan "<path to xlsx>" [--dry-run]
      [--only ID]... [--limit N]
"""
import argparse
import csv
import glob
import hashlib
import json
import os
import re
import subprocess
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import yaml
from PIL import Image, ImageOps

# Windows consoles default to cp1252, which can't print ✓ etc.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parents[1]
INSPIRE = REPO / "content" / "inspire"
LEDGER_PATH = INSPIRE / "publish-ledger.json"
REVIEW_DIR = INSPIRE / "_review"

REGION_CONTINENT = {
    "Americas": None,
    "Asia": "Asia",
    "Atlantic & volcanic islands": "Europe",
    "East & Southern Africa": "Africa",
    "Extreme & Summits (cross-region)": None,
    "Middle East & Gulf": "Asia",
    "Nordics & Baltics": "Europe",
    "North & West Africa": "Africa",
    "Oceania & Pacific": "Oceania",
    "Southern Europe": "Europe",
    "Switzerland & the Alps": "Europe",
}

TRIP_LENGTHS = ["Day", "Weekend", "Week+", "Rally", "Expedition"]
DIFFICULTIES = ["Easy", "Moderate", "Challenging", "Demanding", "Expert"]
ACTIVITIES = [
    "Hiking", "Skiing", "Rafting", "Via Ferrata", "Mountaineering", "Diving",
    "Kayaking", "Bungee", "Seven Summits", "Africa Rally", "Safari", "Roadtrip",
]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DOCX_META_KEYS = {
    "id", "slug", "title", "type", "region", "country", "angle", "guide_wave",
    "links_to_guide", "status", "target_query", "meta_description", "series",
    "source", "pics", "notes", "source_file", "route_file", "longform_file",
    "photos_file", "resolved", "last_reviewed",
}

# Template noise inside story docx files that must never reach the body:
# a literal FRONTMATTER heading and pipe-separated tag lines like
# "Inspire | List | Brand".
NOISE_LINE = re.compile(
    r"^(frontmatter|inspire(\s*\|\s*[\w &+-]+)+|[\w &+-]+(\s*\|\s*[\w &+-]+){1,3})$",
    re.I,
)

CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")


def load_env_local():
    env = REPO / ".env.local"
    if not env.exists():
        return
    for line in env.read_text(encoding="utf-8").splitlines():
        m = re.match(r"^([A-Z_][A-Z0-9_]*)=(.*)$", line.strip())
        if m:
            v = m.group(2).strip().strip('"').strip("'")
            os.environ.setdefault(m.group(1), v)


def find_ffmpeg():
    cand = glob.glob(
        os.path.expandvars(
            r"%LOCALAPPDATA%\Microsoft\WinGet\Packages\Gyan.FFmpeg*\**\bin\ffmpeg.exe"
        ),
        recursive=True,
    )
    if cand:
        return cand[0]
    return "ffmpeg"  # hope it's on PATH


def read_plan(plan_path):
    wb = openpyxl.load_workbook(plan_path, read_only=True, data_only=True)
    ws = wb["Inspire"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[2] is None:
            continue
        rows.append({
            "region": str(r[0] or "").strip(),
            "country": str(r[1] or "").strip(),
            "id": str(r[2]).strip(),
            "title": str(r[3] or "").strip(),
            "angle": str(r[4] or "").strip(),
            "note": str(r[5] or "").strip(),
            "guide_wave": str(r[6] or "").strip(),
            "status": str(r[7] or "").strip(),
            "links_to_guide": str(r[9] or "").strip(),
            "source": str(r[10] or "").strip(),
            "series": str(r[12] or "").strip(),
        })
    wb.close()
    return rows


def find_story_folder(story_id):
    """Story folders live at inspire/{country}/{ID} or, once published and
    filed away, inspire/published/{country}/{ID} — search both depths."""
    hits = sorted(
        h
        for pattern in (INSPIRE / "*" / story_id, INSPIRE / "*" / "*" / story_id)
        for h in glob.glob(str(pattern))
        if Path(h).is_dir()
    )
    if len(hits) > 1:
        print(f"    ! {story_id} found in {len(hits)} places (using first): "
              + ", ".join(str(Path(h).parent.relative_to(INSPIRE)) for h in hits))
    return Path(hits[0]) if hits else None


def extract_docx(docx_path, plan_title):
    """Return (docx_meta: dict, body_markdown: str)."""
    with zipfile.ZipFile(docx_path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
    paras = []
    for pm in re.finditer(r"<w:p[ >].*?</w:p>", xml, re.S):
        block = pm.group(0)
        # <w:t(?:\s...)?> — the attribute part must start with whitespace so
        # self-closing <w:t/> (empty runs) and <w:tab/>/<w:tc> don't match;
        # a loose [^>]* here swallowed raw XML between empty runs and real
        # text into the story body.
        text = "".join(
            m.group(1) for m in re.finditer(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", block, re.S)
        )
        text = (
            text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
            .replace("&apos;", "'").replace("&quot;", '"').strip()
        )
        if text:
            paras.append(text)

    meta = {}
    body = []
    for p in paras:
        # Header region (before the first real paragraph): consume meta
        # keys and drop template noise. Once the body starts, everything
        # is story text.
        if not body:
            km = re.match(r"^([a-z_]+):\s*(.*)$", p)
            if km and km.group(1) in DOCX_META_KEYS:
                meta[km.group(1)] = km.group(2).strip()
                continue
            if NOISE_LINE.match(p.strip()):
                continue
            if plan_title and p.strip().lower() == plan_title.strip().lower():
                continue
        body.append(p)
    return meta, "\n\n".join(body).strip()


def convert_images(folder, story_id, force=False):
    gen = folder / "generated"
    gen.mkdir(exist_ok=True)
    out = []
    photos = sorted((folder / "photos").glob("*"))
    for p in photos:
        if p.suffix.lower() not in (".jpg", ".jpeg", ".png", ".webp"):
            continue
        dst = gen / (p.stem + ".jpg")
        if force or not dst.exists() or dst.stat().st_mtime < p.stat().st_mtime:
            img = Image.open(p)
            # Apply the EXIF orientation tag BEFORE stripping metadata,
            # otherwise phone photos shot in portrait come out rotated.
            img = ImageOps.exif_transpose(img)
            img = img.convert("RGB")
            if img.width > 2560:
                img = img.resize((2560, round(img.height * 2560 / img.width)), Image.LANCZOS)
            img.save(dst, "JPEG", quality=85, optimize=True)
        out.append(dst.name)
    return out


def slot_of(name, story_id=None):
    """Slot number from a sequenced filename ({ID}_3-detail.mp4 -> 3), or None.
    The story-ID prefix is stripped first — IDs like 2022_7-summits-... would
    otherwise match their own digits."""
    if story_id and name.startswith(f"{story_id}_"):
        m = re.match(r"(\d+)[-.]", name[len(story_id) + 1:])
    else:
        m = re.search(r"_(\d+)[-.]", name)
    return int(m.group(1)) if m else None


def source_videos(folder):
    return [
        p for p in sorted((folder / "photos").glob("*"))
        if p.suffix.lower() in (".mp4", ".mov", ".webm")
    ]


def clip_start(name):
    """Cut-start seconds encoded in the video filename label: both
    "..._1-from 5 sec.MOV" and "..._3-From sec 5.MP4" mean start at 5s.
    Default 0.5s when no "from" marker is present."""
    m = re.search(r"from\s*(?:sec\w*\s*)?(\d+(?:\.\d+)?)", name, re.IGNORECASE)
    return float(m.group(1)) if m else 0.5


def cut_clips(folder, story_id, ffmpeg, force=False):
    """7s muted 720p clip per source video, named after its master so the
    _{n}- slot number survives into generated/. Returns [(name, slot), ...]
    in authored order."""
    clips = []
    for vid in source_videos(folder):
        dst = folder / "generated" / (vid.stem + ".mp4")
        if force or not dst.exists() or dst.stat().st_mtime < vid.stat().st_mtime:
            cmd = [
                ffmpeg, "-y", "-v", "error", "-ss", str(clip_start(vid.stem)), "-t", "7", "-i", str(vid),
                "-vf", "scale=-2:720,fps=24", "-c:v", "libx264", "-preset", "slow",
                "-crf", "27", "-pix_fmt", "yuv420p", "-an", "-movflags", "+faststart",
                str(dst),
            ]
            r = subprocess.run(cmd, capture_output=True, text=True)
            if r.returncode != 0:
                print(f"    ! clip failed ({vid.name}): {r.stderr.strip()[:200]}")
                continue
        clips.append((dst.name, slot_of(vid.name, story_id)))
    return clips


def clean_orphan_renditions(folder, story_id):
    """Remove generated/ media whose master no longer exists in photos/, so
    renaming (reorder), replacing, or deleting a master is a pure re-run.
    Also retires the legacy single {ID}_teaser.mp4 (superseded by per-video
    clips). Non-media outputs (story markdown, meta) are never touched."""
    photos_dir = folder / "photos"
    gen = folder / "generated"
    if not photos_dir.is_dir() or not gen.is_dir():
        return []
    expected = {p.stem for p in photos_dir.iterdir() if p.is_file()}
    removed = []
    for f in gen.iterdir():
        if not f.is_file() or f.suffix.lower() not in (".jpg", ".mp4"):
            continue
        if not f.name.startswith(f"{story_id}_"):
            continue
        if f.stem not in expected:
            f.unlink()
            removed.append(f.name)
    return removed


def claude_meta(row, docx_meta, body, photo_names):
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        return {}
    prompt = f"""You draft metadata for a travel story on testedroutes.com. Respond with ONLY a JSON object, no prose.

Story title: {row['title']}
Country: {row['country']} | Region: {row['region']} | Angle: {row['angle']}
Editorial note: {row['note']}
Existing meta description candidate: {docx_meta.get('meta_description', '(none)')}
Target query: {docx_meta.get('target_query', '(none)')}
Photos: {', '.join(photo_names[:6])}

Story text:
{body[:4000]}

Return JSON with exactly these keys:
- "place": the specific place the story happens (e.g. "La Digue", "Darvaza"), or null if none is clearly identifiable. Never just repeat the country.
- "subtitle": one sentence, max 90 chars, factual, in the story's voice. Use en-dash (–) never em-dash.
- "hero_alt": alt text for the hero photo, max 120 chars, descriptive, includes place and country.
- "meta_description": max 160 chars, improves on the candidate if one exists. En-dash only.
- "keywords": 4-6 lowercase search phrases.
- "months": array of months from {MONTHS} when this trip is doable, ONLY if the story or common knowledge makes it clear; else [].
- "trip_length": one of {TRIP_LENGTHS} or null if unclear.
- "difficulty": one of {DIFFICULTIES} or null if not applicable (e.g. essays/lists).
- "activity": array drawn ONLY from {ACTIVITIES}; [] if none apply.

Never invent facts not supported by the text. Prefer null/[] over guessing."""
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps({
            "model": CLAUDE_MODEL,
            "max_tokens": 900,
            "messages": [{"role": "user", "content": prompt}],
        }).encode(),
        headers={
            "x-api-key": key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    for attempt in range(2):
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read())
            text = "".join(b.get("text", "") for b in data.get("content", []))
            jm = re.search(r"\{.*\}", text, re.S)
            return json.loads(jm.group(0)) if jm else {}
        except Exception as e:  # noqa: BLE001
            if attempt == 1:
                print(f"    ! claude meta failed: {e}")
                return {}
    return {}


def prettify(value):
    v = (value or "").strip()
    if not v:
        return v
    if re.fullmatch(r"[a-z0-9-]+", v):
        return " ".join(w.capitalize() for w in v.split("-"))
    return v


def build_meta(row, docx_meta, ai, story_id):
    slug = re.sub(r"^\d{4}_", "", story_id)
    year = re.match(r"^(\d{4})_", story_id)
    meta = {
        "id": story_id,
        "storyId": story_id,
        "slug": slug,
        "title": row["title"],
        "status": "published",
        "language": "en",
        "author": "Paulius Pikelis",
        "destination": prettify(row["country"]),
        "continent": REGION_CONTINENT.get(row["region"]),
        "place": ai.get("place"),
        "subtitle": ai.get("subtitle"),
        "heroAlt": ai.get("hero_alt"),
        "publishedDate": f"{year.group(1)}-01-01" if year else None,
        "date_approximate": bool(year),
        "months": ai.get("months") or [],
        "trip_length": ai.get("trip_length"),
        "difficulty": ai.get("difficulty"),
        "activity": ai.get("activity") or [],
        "seo": {
            "metaTitle": row["title"],
            "metaDescription": ai.get("meta_description") or docx_meta.get("meta_description"),
            "keywords": ai.get("keywords") or [],
        },
        "plan": {
            "region": row["region"],
            "angle": row["angle"],
            "guide_wave": row["guide_wave"],
            "links_to_guide": row["links_to_guide"],
            "series": row["series"] or None,
            "source": row["source"],
            "note": row["note"],
            "target_query": docx_meta.get("target_query"),
        },
    }
    return meta


def content_hash(folder, story_id):
    h = hashlib.md5()
    for name in [f"{story_id}_meta_en.yaml"]:
        p = folder / name
        if p.exists():
            h.update(p.read_bytes())
    gen = folder / "generated"
    if gen.exists():
        for p in sorted(gen.glob("*")):
            h.update(p.name.encode())
            h.update(str(p.stat().st_size).encode())
    return h.hexdigest()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--plan", required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", action="append", default=[])
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--force-media", action="store_true")
    ap.add_argument("--republish", action="store_true",
                    help="also process rows already marked Published")
    args = ap.parse_args()

    load_env_local()
    ffmpeg = find_ffmpeg()

    all_rows = read_plan(args.plan)
    if args.only:
        # Explicit IDs override the status filter (used for targeted
        # publishes the user requested directly).
        rows = [r for r in all_rows if r["id"] in args.only]
    else:
        wanted = {"to be published"} | ({"published"} if args.republish else set())
        rows = [r for r in all_rows if r["status"].lower() in wanted]
    if args.limit:
        rows = rows[: args.limit]
    print(f"stories to publish: {len(rows)}  (dry-run={args.dry_run})")

    ledger = json.loads(LEDGER_PATH.read_text()) if LEDGER_PATH.exists() else {}
    published, failed = [], []
    review_rows = []

    for i, row in enumerate(rows, 1):
        sid = row["id"]
        print(f"\n[{i}/{len(rows)}] {sid}")
        folder = find_story_folder(sid)
        if not folder:
            print("    ! folder not found"); failed.append((sid, "folder not found")); continue

        docx_candidates = sorted(folder.glob("*story*.docx")) + sorted(folder.glob("*story*.md"))
        if not docx_candidates:
            print("    ! no story file"); failed.append((sid, "no story file")); continue

        try:
            src = docx_candidates[0]
            if src.suffix == ".docx":
                docx_meta, body = extract_docx(src, row["title"])
            else:
                docx_meta, body = {}, src.read_text(encoding="utf-8")
            if len(body) < 200:
                print(f"    ! story too short ({len(body)} chars)"); failed.append((sid, "story too short")); continue

            gen = folder / "generated"
            gen.mkdir(exist_ok=True)
            (gen / f"{sid}_story_en.md").write_text(body + "\n", encoding="utf-8")

            photos = convert_images(folder, sid, force=args.force_media)
            clips = cut_clips(folder, sid, ffmpeg, force=args.force_media)
            removed = clean_orphan_renditions(folder, sid)
            slot = clips[0][1] if clips else None
            clip_note = f", {len(clips)} video clips" if clips else ""
            orphan_note = f", {len(removed)} stale renditions removed" if removed else ""
            print(f"    media: {len(photos)} photos{clip_note}{orphan_note}")

            meta_path = folder / f"{sid}_meta_en.yaml"
            if meta_path.exists():
                meta = yaml.safe_load(meta_path.read_text(encoding="utf-8"))
                # backfill the video slot into older meta files
                if slot and meta.get("video_slot") != slot:
                    meta["video_slot"] = slot
                    meta_path.write_text(
                        yaml.safe_dump(meta, sort_keys=False, allow_unicode=True, width=100),
                        encoding="utf-8",
                    )
                print("    meta: existing file kept")
            else:
                ai = claude_meta(row, docx_meta, body, photos)
                meta = build_meta(row, docx_meta, ai, sid)
                if slot:
                    meta["video_slot"] = slot
                meta_path.write_text(
                    yaml.safe_dump(meta, sort_keys=False, allow_unicode=True, width=100),
                    encoding="utf-8",
                )
                print(f"    meta: drafted (place={meta.get('place')}, trip_length={meta.get('trip_length')})")

            cmd = ["node", "--env-file=.env.local", "scripts/publish-to-sanity.mjs", str(folder)]
            if args.dry_run:
                cmd.append("--dry-run")
            r = subprocess.run(
                cmd, capture_output=True, text=True, cwd=REPO,
                encoding="utf-8", errors="replace",
            )
            ok = r.returncode == 0 and ("✓ Published" in r.stdout or "Dry run complete" in r.stdout)
            if not ok:
                tail = (r.stdout + r.stderr).strip().splitlines()[-8:]
                print("    ! publish failed:\n      " + "\n      ".join(tail))
                failed.append((sid, "publish failed"))
                continue

            print("    ✓ published" if not args.dry_run else "    ✓ dry-run ok")
            published.append(sid)
            review_rows.append({
                "id": sid,
                "title": row["title"],
                "place": meta.get("place"),
                "subtitle": meta.get("subtitle"),
                "meta_description": (meta.get("seo") or {}).get("metaDescription"),
                "months": ",".join(meta.get("months") or []),
                "trip_length": meta.get("trip_length"),
                "difficulty": meta.get("difficulty"),
                "activity": ",".join(meta.get("activity") or []),
                "url": f"https://testedroutes.com/inspire/{meta.get('slug')}",
            })
            if not args.dry_run:
                ledger[sid] = {
                    "slug": meta.get("slug"),
                    "publishedAt": datetime.now(timezone.utc).isoformat(),
                    "hash": content_hash(folder, sid),
                    "langs": ["en"],
                }
        except Exception as e:  # noqa: BLE001
            print(f"    ! error: {e}")
            failed.append((sid, str(e)))

    if not args.dry_run and published:
        LEDGER_PATH.write_text(json.dumps(ledger, indent=1), encoding="utf-8")

    if review_rows:
        REVIEW_DIR.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        rp = REVIEW_DIR / f"publish-review-{stamp}.csv"
        with rp.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(review_rows[0].keys()))
            w.writeheader()
            w.writerows(review_rows)
        print(f"\nreview sheet: {rp}")

    # masterfile write-back: only on real publishes
    if not args.dry_run and published:
        try:
            wb = openpyxl.load_workbook(args.plan)
            ws = wb["Inspire"]
            done = set(published)
            for r in ws.iter_rows(min_row=2):
                if r[2].value and str(r[2].value).strip() in done:
                    r[7].value = "Published"
            wb.save(args.plan)
            print(f"masterfile updated: {len(done)} rows -> Published")
        except PermissionError:
            alt = str(args.plan).replace(".xlsx", "_PUBLISHED-UPDATE.xlsx")
            wb.save(alt)
            print(f"! masterfile locked (open in Excel?) - wrote {alt} instead")

    print(f"\ndone: {len(published)} published, {len(failed)} failed")
    for sid, why in failed:
        print(f"  FAILED {sid}: {why}")


if __name__ == "__main__":
    main()
