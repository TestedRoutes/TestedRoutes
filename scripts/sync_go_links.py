#!/usr/bin/env python3
"""
sync_go_links.py - Country master xlsx -> Sanity affiliateLink docs.

Every QR printed in a guide encodes  testedroutes.com/go/<slug>.  The /go/
route resolves that slug against the affiliateLink collection in Sanity
FIRST (see app/go/[slug]/route.js), so a slug with no doc silently falls
through to the homepage. This script creates the missing docs from the
country master, which is where the author maintains the destinations.

Sanity is deliberately the store rather than the CURATED dict in code:
destinations must be re-pointable after PDFs are sold, without a deploy.

Booking links are pre-baked URLs (Google Maps, hotel sites, GetYourGuide),
so program is always "other" = pass-through, no tracking param injected.
Swap program later per-link in Studio if a destination becomes affiliate.

Usage:
  python scripts/sync_go_links.py --file "<master.xlsx>"              # dry run
  python scripts/sync_go_links.py --file "<master.xlsx>" --live       # write
  python scripts/sync_go_links.py --file "<master.xlsx>" --sheet "X"

Env (.env.local): NEXT_PUBLIC_SANITY_PROJECT_ID, NEXT_PUBLIC_SANITY_DATASET,
                  SANITY_API_WRITE_TOKEN
"""
import argparse
import json
import os
import re
import sys
import urllib.request
import urllib.error

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

API_VERSION = "2026-04-24"
REDIRECT_HOST = "testedroutes.com/go/"

# Excel Type -> affiliateLink.category (the schema's fixed list).
TYPE_TO_CATEGORY = {
    "restaurant": "restaurant",
    "stay": "hotels",
    "sleep": "hotels",
    "activity": "tickets",
    "activity / tour": "tickets",
    "beach access": "tickets",
    "sight": "tickets",
    "viewpoint": "tickets",
    "beach": "tickets",
    "car rental": "cars",
    "car": "cars",
    "ferry": "transport",
    "logistics": "transport",
    "paperwork": "tickets",
}
# Travel essentials are recognised by name, not type.
NAME_TO_CATEGORY = [
    ("revolut", "currency"),
    ("wise", "currency"),
    ("nordvpn", "vpn"),
    ("vpn", "vpn"),
    ("esim", "esim"),
    ("saily", "esim"),
    ("holafly", "esim"),
    ("insurance", "insurance"),
]

SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9-]*$")


def load_env(path=".env.local"):
    if not os.path.exists(path):
        sys.exit(f"{path} not found - run from the repo root")
    for line in open(path, encoding="utf-8"):
        m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)=(.*)$", line)
        if m:
            os.environ.setdefault(m.group(1), m.group(2).strip().strip('"'))


def sanity(path, token, project, dataset, method="GET", body=None):
    url = f"https://{project}.api.sanity.io/v{API_VERSION}/{path}"
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(
        url, data=data, method=method,
        headers={"Authorization": f"Bearer {token}",
                 "Content-Type": "application/json",
                 "User-Agent": "TestedRoutes-sync-go-links/1.0"},
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read())


def slug_from(shortlink):
    """testedroutes.com/go/sey-x -> sey-x  (tolerates scheme + trailing /)"""
    s = str(shortlink or "").strip().rstrip("/")
    if not s:
        return None
    s = re.sub(r"^https?://", "", s, flags=re.I)
    i = s.lower().find(REDIRECT_HOST)
    if i >= 0:
        s = s[i + len(REDIRECT_HOST):]
    elif "/" in s:
        s = s.rsplit("/", 1)[1]
    return s.lower().strip() or None


def category_for(name, typ):
    low = str(name or "").lower()
    for needle, cat in NAME_TO_CATEGORY:
        if needle in low:
            return cat
    return TYPE_TO_CATEGORY.get(str(typ or "").strip().lower(), "tickets")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet")
    ap.add_argument("--live", action="store_true")
    ap.add_argument("--scope", default="guide", choices=["guide", "global"])
    a = ap.parse_args()

    load_env()
    project = os.environ.get("NEXT_PUBLIC_SANITY_PROJECT_ID")
    dataset = os.environ.get("NEXT_PUBLIC_SANITY_DATASET", "production")
    token = os.environ.get("SANITY_API_WRITE_TOKEN")
    if not project or not token:
        sys.exit("NEXT_PUBLIC_SANITY_PROJECT_ID and SANITY_API_WRITE_TOKEN are required")

    wb = load_workbook(a.file, read_only=True, data_only=True)
    ws = wb[a.sheet] if a.sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[0]]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    ci = {
        "name": col("Name"),
        "type": col("Type"),
        "qr":   col("QR shortlink (go/)", "QR shortlink"),
        "url":  col("Booking link (destination)", "Booking link"),
    }
    missing = [k for k, v in ci.items() if v is None]
    if missing:
        sys.exit(f"missing column(s) for: {missing}\nfound: {hdr}")

    planned, problems, seen = [], [], {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        name = str(r[ci["name"]] or "").strip()
        qr   = r[ci["qr"]]
        url  = str(r[ci["url"]] or "").strip()
        typ  = r[ci["type"]]

        slug = slug_from(qr)
        if not slug or slug.upper() in ("N/A", "NA", "NONE"):
            continue                                   # no QR printed - skip
        if not SLUG_RE.match(slug):
            problems.append((name, slug, "slug not lowercase/hyphenated"))
            continue
        if not url or url.upper() in ("N/A", "NA", "NONE"):
            problems.append((name, slug, "QR printed but NO booking link -> would 302 to homepage"))
            continue
        if not re.match(r"^https?://", url, re.I):
            problems.append((name, slug, f"destination is not a URL: {url[:48]}"))
            continue
        if slug in seen:
            problems.append((name, slug, f"duplicate slug (also '{seen[slug]}')"))
            continue
        seen[slug] = name
        planned.append({"slug": slug, "label": name[:120], "url": url,
                        "category": category_for(name, typ)})

    print(f"\nsync_go_links  •  {os.path.basename(a.file)}  •  "
          f"{'LIVE' if a.live else 'DRY RUN'}\n")
    print(f"{len(planned)} link(s) ready, {len(problems)} problem(s)\n")

    existing = sanity(
        f"data/query/{dataset}?query=" + urllib.parse.quote(
            '*[_type=="affiliateLink"]{_id,"slug":slug.current,url}'),
        token, project, dataset)["result"]
    by_slug = {e["slug"]: e for e in existing if e.get("slug")}

    creates, updates, unchanged = [], [], 0
    for p in planned:
        cur = by_slug.get(p["slug"])
        if not cur:
            creates.append(p)
        elif (cur.get("url") or "") != p["url"]:
            updates.append((p, cur))
        else:
            unchanged += 1

    for p in creates:
        print(f"  + create  {p['slug']:34} {p['category']:11} -> {p['url'][:58]}")
    for p, cur in updates:
        print(f"  ~ update  {p['slug']:34} {cur.get('url','')[:30]} -> {p['url'][:40]}")
    if unchanged:
        print(f"  = {unchanged} already correct")

    if problems:
        print("\nPROBLEMS (no doc written for these):")
        for name, slug, why in problems:
            print(f"  ! {str(name)[:34]:36} {str(slug):26} {why}")

    if not a.live:
        print("\nDry run - nothing written. Re-run with --live to apply.")
        return

    muts = []
    for p in creates:
        muts.append({"create": {
            "_type": "affiliateLink", "label": p["label"],
            "slug": {"_type": "slug", "current": p["slug"]},
            "scope": a.scope, "category": p["category"],
            "url": p["url"], "program": "other",
            "notes": f"Synced from {os.path.basename(a.file)} by sync_go_links.py.",
        }})
    for p, cur in updates:
        muts.append({"patch": {"id": cur["_id"], "set": {"url": p["url"], "label": p["label"]}}})

    if not muts:
        print("\nNothing to write.")
        return
    res = sanity(f"data/mutate/{dataset}", token, project, dataset,
                 method="POST", body={"mutations": muts})
    print(f"\nWrote {len(res.get('results', []))} document(s).")


if __name__ == "__main__":
    import urllib.parse
    main()
